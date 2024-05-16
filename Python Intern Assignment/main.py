import os
import re
import PyPDF2
from docx import Document
import openpyxl

def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ''
        for page_num in range(len(reader.pages)):
            text += reader.pages[page_num].extract_text()
    return text

def extract_text_from_docx(docx_path):
    doc = Document(docx_path)
    text = ''
    for para in doc.paragraphs:
        text += para.text + '\n'
    return text

def extract_info(text):
    email_pattern = r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'
    phone_pattern = r'(\+\d{1,2}\s)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}'
    
    emails = re.findall(email_pattern, text)
    phones = re.findall(phone_pattern, text)
    
    return emails, phones, text

def create_excel(emails, phones, texts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CV Data"
    ws['A1'] = 'Email ID'
    ws['B1'] = 'Contact No.'
    ws['C1'] = 'Text'
    
    for row_num, (email, phone, text) in enumerate(zip(emails, phones, texts), start=2):
        ws[f'A{row_num}'] = email
        ws[f'B{row_num}'] = phone
        ws[f'C{row_num}'] = text
    
    excel_path = 'cv_data.xlsx'
    wb.save(excel_path)
    return excel_path

def process_cv_files(directory_path):
    emails = []
    phones = []
    texts = []
    
    for filename in os.listdir(directory_path):
        if filename.endswith('.pdf'):
            text = extract_text_from_pdf(os.path.join(directory_path, filename))
        elif filename.endswith('.docx'):
            text = extract_text_from_docx(os.path.join(directory_path, filename))
        else:
            continue
        
        extracted_emails, extracted_phones, extracted_text = extract_info(text)
        
        emails.extend(extracted_emails)
        phones.extend(extracted_phones)
        texts.append(extracted_text)
    
    excel_path = create_excel(emails, phones, texts)
    return excel_path

if __name__ == '__main__':
    directory_path = input("Enter the directory path containing CVs: ")
    excel_file = process_cv_files(directory_path)
    print(f"Excel file saved at: {excel_file}")
